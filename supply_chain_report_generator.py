"""
Supply Chain Report Generator — produces XLSX and CSV for SC Intelligence reports.

Same pattern as report_generator.py but with supply-chain-specific sections.
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

# ─── Color palette (dark with orange accent) ───
DARK_BG = "0A0A0A"
SECTION_BG = "1A1A1A"
HEADER_BG = "262626"
ACCENT = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "E0E0E0"
TEXT_DARK = "1A1A1A"
TEXT_MED = "444444"

# Risk level colors
RISK_CRITICAL = "EF4444"
RISK_HIGH = "F97316"
RISK_MEDIUM = "EAB308"
RISK_LOW = "22C55E"


def _safe(val, default="Not available"):
    if val is None:
        return default
    if isinstance(val, str):
        return val.strip() or default
    return str(val)


def _risk_color(level: str) -> str:
    """Return hex color for risk level."""
    level = (level or "").lower()
    if "critical" in level or level in ("9", "10"):
        return RISK_CRITICAL
    if "high" in level or level in ("7", "8"):
        return RISK_HIGH
    if "medium" in level or "moderate" in level or level in ("5", "6"):
        return RISK_MEDIUM
    return RISK_LOW


def generate_sc_csv(report: dict, company_name: str) -> str:
    """Generate CSV for supply chain intelligence report."""
    if not report or "error" in report:
        return ""

    path = os.path.join(REPORTS_DIR, f"{company_name.replace(' ', '_')}_SC_Intelligence.csv")

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)

        # Executive Summary
        es = report.get("executive_summary", {})
        w.writerow(["EXECUTIVE SUMMARY", "", "", "", "", "", ""])
        w.writerow(["Company Name", _safe(es.get("company_name")), "", "", "", "", ""])
        w.writerow(["Industry", _safe(es.get("industry")), "", "", "", "", ""])
        w.writerow(["Supply Chain Complexity", _safe(es.get("supply_chain_complexity")), "", "", "", "", ""])
        w.writerow(["Overall Risk Score", _safe(es.get("overall_risk_score")), "", "", "", "", ""])
        w.writerow(["Top Findings", _safe(es.get("top_findings")), "", "", "", "", ""])
        w.writerow(["Key Recommendation", _safe(es.get("key_recommendation")), "", "", "", "", ""])
        w.writerow([""])

        # Supplier Map
        w.writerow(["SUPPLIER MAP", "", "", "", "", "", ""])
        w.writerow(["Tier", "Supplier", "Location", "Product Category", "Est. Revenue Share", "Risk Score", "Alternatives"])
        for s in report.get("supplier_map", []):
            w.writerow([
                _safe(s.get("tier")), _safe(s.get("supplier")),
                _safe(s.get("location")), _safe(s.get("product_category")),
                _safe(s.get("est_revenue_share")), _safe(s.get("risk_score")),
                _safe(s.get("alternatives_available")),
            ])
        w.writerow([""])

        # Geographic Risk
        gr = report.get("geographic_risk", {})
        w.writerow(["GEOGRAPHIC RISK", "", "", "", "", "", ""])
        w.writerow(["Concentration Index (HHI)", _safe(gr.get("concentration_index")), "", "", "", "", ""])
        w.writerow(["Primary Regions", _safe(gr.get("primary_regions")), "", "", "", "", ""])
        w.writerow(["Natural Disaster Exposure", _safe(gr.get("natural_disaster_exposure")), "", "", "", "", ""])
        w.writerow(["Political Stability Score", _safe(gr.get("political_stability_score")), "", "", "", "", ""])
        w.writerow(["Logistics Reliability", _safe(gr.get("logistics_reliability")), "", "", "", "", ""])
        w.writerow([""])

        # Dependency Analysis
        w.writerow(["DEPENDENCY ANALYSIS", "", "", "", "", "", ""])
        w.writerow(["Vulnerability", "Severity", "Description", "Mitigation", "", "", ""])
        for d in report.get("dependency_analysis", []):
            w.writerow([
                _safe(d.get("vulnerability")), _safe(d.get("severity")),
                _safe(d.get("description")), _safe(d.get("mitigation")),
                "", "", "",
            ])
        w.writerow([""])

        # Logistics Infrastructure
        li = report.get("logistics_infrastructure", {})
        w.writerow(["LOGISTICS INFRASTRUCTURE", "", "", "", "", "", ""])
        w.writerow(["Primary Transport Modes", _safe(li.get("primary_transport_modes")), "", "", "", "", ""])
        w.writerow(["Key Chokepoints", _safe(li.get("key_chokepoints")), "", "", "", "", ""])
        w.writerow(["Chicago Hub Dependency", _safe(li.get("chicago_hub_dependency")), "", "", "", "", ""])
        w.writerow(["Lead Time Estimate", _safe(li.get("lead_time_estimate")), "", "", "", "", ""])
        w.writerow(["Freight Cost Trend", _safe(li.get("freight_cost_trend")), "", "", "", "", ""])
        w.writerow([""])

        # Trade & Regulatory Risk
        tr = report.get("trade_regulatory_risk", {})
        w.writerow(["TRADE & REGULATORY RISK", "", "", "", "", "", ""])
        w.writerow(["Tariff Exposure", _safe(tr.get("tariff_exposure")), "", "", "", "", ""])
        w.writerow(["Sanctions Screening", _safe(tr.get("sanctions_screening")), "", "", "", "", ""])
        w.writerow(["Trade Policy Sensitivity", _safe(tr.get("trade_policy_sensitivity")), "", "", "", "", ""])
        w.writerow(["Estimated Tariff Impact", _safe(tr.get("estimated_tariff_impact")), "", "", "", "", ""])
        w.writerow([""])

        # Competitive Displacement
        w.writerow(["COMPETITIVE DISPLACEMENT", "", "", "", "", "", ""])
        w.writerow(["Signal", "Competitor", "Impact", "Confidence", "", "", ""])
        for c in report.get("competitive_displacement", []):
            w.writerow([
                _safe(c.get("signal")), _safe(c.get("competitor")),
                _safe(c.get("impact")), _safe(c.get("confidence")),
                "", "", "",
            ])
        w.writerow([""])

        # Scenarios
        w.writerow(["SCENARIO ANALYSIS", "", "", "", "", "", ""])
        w.writerow(["Scenario", "Probability", "Impact", "Response Recommendation", "", "", ""])
        for s in report.get("scenarios", []):
            w.writerow([
                _safe(s.get("scenario")), _safe(s.get("probability")),
                _safe(s.get("impact")), _safe(s.get("response_recommendation")),
                "", "", "",
            ])
        w.writerow([""])

        # Recommendations
        rec = report.get("recommendations", {})
        w.writerow(["RECOMMENDATIONS", "", "", "", "", "", ""])
        w.writerow(["Immediate Actions", _safe(rec.get("immediate_actions")), "", "", "", "", ""])
        w.writerow(["Monitoring Indicators", _safe(rec.get("monitoring_indicators")), "", "", "", "", ""])
        w.writerow(["Further Investigation", _safe(rec.get("further_investigation")), "", "", "", "", ""])

    print(f"  📄 CSV: {path}")
    return path


def generate_sc_xlsx(report: dict, company_name: str) -> str:
    """Generate professionally formatted XLSX for supply chain intelligence."""
    if not report or "error" in report:
        return ""

    path = os.path.join(REPORTS_DIR, f"{company_name.replace(' ', '_')}_SC_Intelligence.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Supply Chain Intelligence"

    # Column widths
    for col, width in [("A", 28), ("B", 32), ("C", 38), ("D", 30), ("E", 25), ("F", 18), ("G", 22)]:
        ws.column_dimensions[col].width = width

    # Styles
    section_font = Font(name="Arial", size=13, bold=True, color=WHITE)
    section_fill = PatternFill("solid", fgColor=DARK_BG)
    header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    label_font = Font(name="Arial", size=10, bold=True, color=TEXT_DARK)
    value_font = Font(name="Arial", size=10, color=TEXT_MED)
    accent_font = Font(name="Arial", size=12, bold=True, color=ACCENT)
    thin_border = Border(bottom=Side(style="thin", color=MED_GRAY))
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    row = 1

    def write_section_header(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 32
        row += 1

    def write_table_header(cols):
        nonlocal row
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        ws.row_dimensions[row].height = 24
        row += 1

    def write_kv(label, value):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.alignment = wrap
        c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=_safe(value))
        c2.font = value_font
        c2.alignment = wrap
        c2.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

    def write_data_row(values, alt=False):
        nonlocal row
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if alt else PatternFill()
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=_safe(val) if val else "")
            cell.font = value_font
            cell.alignment = wrap
            cell.border = thin_border
            if alt:
                cell.fill = fill
        ws.row_dimensions[row].height = 22
        row += 1

    def write_spacer():
        nonlocal row
        ws.row_dimensions[row].height = 12
        row += 1

    # ── EXECUTIVE SUMMARY ──
    write_section_header("EXECUTIVE SUMMARY")
    es = report.get("executive_summary", {})
    write_kv("Company Name", es.get("company_name"))
    write_kv("Industry", es.get("industry"))
    write_kv("Supply Chain Complexity", es.get("supply_chain_complexity"))

    # Risk score with color
    c1 = ws.cell(row=row, column=1, value="Overall Risk Score")
    c1.font = label_font
    c1.border = thin_border
    risk_val = _safe(es.get("overall_risk_score"))
    c2 = ws.cell(row=row, column=2, value=risk_val)
    c2.font = Font(name="Arial", size=14, bold=True, color=_risk_color(risk_val))
    c2.border = thin_border
    row += 1

    write_kv("Top Findings", es.get("top_findings"))
    write_kv("Key Recommendation", es.get("key_recommendation"))
    write_spacer()

    # ── SUPPLIER MAP ──
    write_section_header("SUPPLIER MAP")
    write_table_header(["Tier", "Supplier", "Location", "Product Category", "Revenue Share", "Risk Score", "Alternatives"])
    for i, s in enumerate(report.get("supplier_map", [])):
        write_data_row([
            s.get("tier"), s.get("supplier"), s.get("location"),
            s.get("product_category"), s.get("est_revenue_share"),
            s.get("risk_score"), s.get("alternatives_available"),
        ], alt=i % 2 == 1)
    write_spacer()

    # ── GEOGRAPHIC RISK ──
    write_section_header("GEOGRAPHIC RISK")
    gr = report.get("geographic_risk", {})
    write_kv("Concentration Index (HHI)", gr.get("concentration_index"))
    write_kv("Primary Regions", gr.get("primary_regions"))
    write_kv("Natural Disaster Exposure", gr.get("natural_disaster_exposure"))
    write_kv("Political Stability Score", gr.get("political_stability_score"))
    write_kv("Logistics Reliability", gr.get("logistics_reliability"))
    write_spacer()

    # ── DEPENDENCY ANALYSIS ──
    write_section_header("DEPENDENCY ANALYSIS")
    write_table_header(["Vulnerability", "Severity", "Description", "Mitigation", "", "", ""])
    for i, d in enumerate(report.get("dependency_analysis", [])):
        write_data_row([
            d.get("vulnerability"), d.get("severity"),
            d.get("description"), d.get("mitigation"),
            None, None, None,
        ], alt=i % 2 == 1)
    write_spacer()

    # ── LOGISTICS INFRASTRUCTURE ──
    write_section_header("LOGISTICS INFRASTRUCTURE")
    li = report.get("logistics_infrastructure", {})
    write_kv("Primary Transport Modes", li.get("primary_transport_modes"))
    write_kv("Key Chokepoints", li.get("key_chokepoints"))
    write_kv("Chicago Hub Dependency", li.get("chicago_hub_dependency"))
    write_kv("Lead Time Estimate", li.get("lead_time_estimate"))
    write_kv("Freight Cost Trend", li.get("freight_cost_trend"))
    write_spacer()

    # ── TRADE & REGULATORY RISK ──
    write_section_header("TRADE & REGULATORY RISK")
    tr = report.get("trade_regulatory_risk", {})
    write_kv("Tariff Exposure", tr.get("tariff_exposure"))
    write_kv("Sanctions Screening", tr.get("sanctions_screening"))
    write_kv("Trade Policy Sensitivity", tr.get("trade_policy_sensitivity"))
    write_kv("Estimated Tariff Impact", tr.get("estimated_tariff_impact"))
    write_spacer()

    # ── COMPETITIVE DISPLACEMENT ──
    write_section_header("COMPETITIVE DISPLACEMENT")
    write_table_header(["Signal", "Competitor", "Impact", "Confidence", "", "", ""])
    for i, c in enumerate(report.get("competitive_displacement", [])):
        write_data_row([
            c.get("signal"), c.get("competitor"),
            c.get("impact"), c.get("confidence"),
            None, None, None,
        ], alt=i % 2 == 1)
    write_spacer()

    # ── SCENARIO ANALYSIS ──
    write_section_header("SCENARIO ANALYSIS")
    write_table_header(["Scenario", "Probability", "Impact", "Response", "", "", ""])
    for i, s in enumerate(report.get("scenarios", [])):
        write_data_row([
            s.get("scenario"), s.get("probability"),
            s.get("impact"), s.get("response_recommendation"),
            None, None, None,
        ], alt=i % 2 == 1)
    write_spacer()

    # ── RECOMMENDATIONS ──
    write_section_header("RECOMMENDATIONS")
    rec = report.get("recommendations", {})
    write_kv("Immediate Actions", rec.get("immediate_actions"))
    write_kv("Monitoring Indicators", rec.get("monitoring_indicators"))
    write_kv("Further Investigation", rec.get("further_investigation"))

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(path)
    print(f"  📊 XLSX: {path}")
    return path


if __name__ == "__main__":
    import json, sys
    sample = json.load(open(sys.argv[1])) if len(sys.argv) > 1 else {}
    if sample:
        generate_sc_xlsx(sample, "Test_Company")
        generate_sc_csv(sample, "Test_Company")
