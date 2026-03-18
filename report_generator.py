"""
Report Generator — produces XLSX and CSV matching the Cofounder.co VC Analysis format.

The output format is reverse-engineered from the reference CSV:
10 sections, 5-column layout, section headers as row separators.
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

# ─── Color palette ───
DARK_BG = "1A1A2E"
SECTION_BG = "16213E"
HEADER_BG = "0F3460"
ACCENT = "E94560"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "E0E0E0"
TEXT_DARK = "1A1A2E"
TEXT_MED = "444444"


def _safe(val, default="Not available"):
    if val is None:
        return default
    if isinstance(val, str):
        return val.strip() or default
    return str(val)


def generate_csv(report: dict, company_name: str) -> str:
    """Generate CSV matching the exact Cofounder.co VC Analysis format."""
    if not report or "error" in report:
        return ""

    path = os.path.join(REPORTS_DIR, f"{company_name.replace(' ', '_')}_VC_Analysis.csv")

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)

        es = report.get("executive_summary", {})
        w.writerow(["EXECUTIVE SUMMARY", "", "", "", ""])
        w.writerow(["Company Name", _safe(es.get("company_name")), "", "", ""])
        w.writerow(["Product", _safe(es.get("product")), "", "", ""])
        w.writerow(["Founded", _safe(es.get("founded")), "", "", ""])
        w.writerow(["Funding Stage", _safe(es.get("funding_stage")), "", "", ""])
        w.writerow(["Lead Investors", _safe(es.get("lead_investors")), "", "", ""])
        w.writerow(["", "", "", "", ""])

        w.writerow(["LEADERSHIP TEAM", "", "", "", ""])
        w.writerow(["Position", "Name", "Background", "LinkedIn", "Key Experience"])
        for person in report.get("leadership_team", []):
            w.writerow([
                _safe(person.get("position")),
                _safe(person.get("name")),
                _safe(person.get("background")),
                _safe(person.get("linkedin")),
                _safe(person.get("key_experience")),
            ])
        ts = _safe(report.get("team_size", ""))
        tb = _safe(report.get("team_breakdown", ""))
        w.writerow(["Team Size", ts, tb, "", ""])
        w.writerow(["", "", "", "", ""])

        w.writerow(["MARKET ANALYSIS", "", "", "", ""])
        w.writerow(["Category", "2024 Size", "2030-2034 Projection", "CAGR", "Notes"])
        for m in report.get("market_analysis", []):
            w.writerow([
                _safe(m.get("category")),
                _safe(m.get("current_size")),
                _safe(m.get("projection")),
                _safe(m.get("cagr")),
                _safe(m.get("notes")),
            ])
        w.writerow(["", "", "", "", ""])

        pp = report.get("product_positioning", {})
        w.writerow(["PRODUCT & POSITIONING", "", "", "", ""])
        w.writerow(["Core Value Prop", _safe(pp.get("core_value_prop")), "", "", ""])
        w.writerow(["Key Differentiators", _safe(pp.get("key_differentiators")), "", "", ""])
        w.writerow(["Target Customers", _safe(pp.get("target_customers")), "", "", ""])
        w.writerow(["Pricing Model", _safe(pp.get("pricing_model")), "", "", ""])
        w.writerow(["", "", "", "", ""])

        w.writerow(["TRACTION METRICS", "", "", "", ""])
        w.writerow(["Metric", "Current Value", "Significance", "", ""])
        for t in report.get("traction_metrics", []):
            w.writerow([
                _safe(t.get("metric")),
                _safe(t.get("current_value")),
                _safe(t.get("significance")),
                "", "",
            ])
        w.writerow(["", "", "", "", ""])

        w.writerow(["COMPETITIVE LANDSCAPE", "", "", "", ""])
        w.writerow(["Competitor", "Valuation/Funding", "Focus", "Strengths", "Weaknesses vs Target"])
        for c in report.get("competitive_landscape", []):
            w.writerow([
                _safe(c.get("competitor")),
                _safe(c.get("valuation_funding")),
                _safe(c.get("focus")),
                _safe(c.get("strengths")),
                _safe(c.get("weaknesses_vs_target")),
            ])
        w.writerow(["", "", "", "", ""])

        fa = report.get("financial_analysis", {})
        w.writerow(["FINANCIAL ANALYSIS", "", "", "", ""])
        w.writerow(["Funding Raised", _safe(fa.get("funding_raised")), "", "", ""])
        w.writerow(["Investors", _safe(fa.get("investors")), "", "", ""])
        w.writerow(["Revenue Model", _safe(fa.get("revenue_model")), "", "", ""])
        w.writerow(["Burn Rate", _safe(fa.get("burn_rate")), "", "", ""])
        w.writerow(["Runway", _safe(fa.get("runway")), "", "", ""])
        w.writerow(["", "", "", "", ""])

        w.writerow(["RISK ASSESSMENT", "", "", "", ""])
        w.writerow(["Risk Category", "Level", "Description", "Mitigation", ""])
        for r in report.get("risk_assessment", []):
            w.writerow([
                _safe(r.get("risk_category")),
                _safe(r.get("level")),
                _safe(r.get("description")),
                _safe(r.get("mitigation")),
                "",
            ])
        w.writerow(["", "", "", "", ""])

        ir = report.get("investment_recommendation", {})
        w.writerow(["INVESTMENT RECOMMENDATION", "", "", "", ""])
        w.writerow(["Overall Rating", _safe(ir.get("overall_rating")), "", "", ""])
        w.writerow(["Investment Thesis", _safe(ir.get("investment_thesis")), "", "", ""])
        w.writerow(["Key Strengths", _safe(ir.get("key_strengths")), "", "", ""])
        w.writerow(["Key Concerns", _safe(ir.get("key_concerns")), "", "", ""])
        w.writerow(["Valuation Assessment", _safe(ir.get("valuation_assessment")), "", "", ""])
        w.writerow(["Recommended Investment", _safe(ir.get("recommended_investment")), "", "", ""])
        w.writerow(["Expected Timeline to Exit", _safe(ir.get("expected_timeline_to_exit")), "", "", ""])
        w.writerow(["Risk-Adjusted Return", _safe(ir.get("risk_adjusted_return")), "", "", ""])

    print(f"  📄 CSV: {path}")
    return path


def generate_xlsx(report: dict, company_name: str) -> str:
    """Generate a professionally formatted XLSX matching the VC Analysis format."""
    if not report or "error" in report:
        return ""

    path = os.path.join(REPORTS_DIR, f"{company_name.replace(' ', '_')}_VC_Analysis.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 32

    # Styles
    section_font = Font(name="Arial", size=13, bold=True, color=WHITE)
    section_fill = PatternFill("solid", fgColor=DARK_BG)
    header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    label_font = Font(name="Arial", size=10, bold=True, color=TEXT_DARK)
    value_font = Font(name="Arial", size=10, color=TEXT_MED)
    accent_font = Font(name="Arial", size=12, bold=True, color=ACCENT)
    thin_border = Border(
        bottom=Side(style="thin", color=MED_GRAY),
    )
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    wrap_bold = Alignment(vertical="top", wrap_text=True)

    row = 1

    def write_section_header(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 32
        row += 1

    def write_table_header(cols):
        nonlocal row
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        ws.row_dimensions[row].height = 24
        row += 1

    def write_kv(label, value):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.alignment = wrap_bold
        c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=_safe(value))
        c2.font = value_font
        c2.alignment = wrap
        c2.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

    def write_data_row(values, alt=False):
        nonlocal row
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if alt else PatternFill()
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=_safe(val) if val else "")
            cell.font = value_font
            cell.alignment = wrap
            cell.border = thin_border
            if alt:
                cell.fill = fill
        ws.row_dimensions[row].height = 22
        row += 1

    def write_spacer():
        nonlocal row
        ws.row_dimensions[row].height = 12
        row += 1

    # ── EXECUTIVE SUMMARY ──
    write_section_header("EXECUTIVE SUMMARY")
    es = report.get("executive_summary", {})
    write_kv("Company Name", es.get("company_name"))
    write_kv("Product", es.get("product"))
    write_kv("Founded", es.get("founded"))
    write_kv("Funding Stage", es.get("funding_stage"))
    write_kv("Lead Investors", es.get("lead_investors"))
    write_spacer()

    # ── LEADERSHIP TEAM ──
    write_section_header("LEADERSHIP TEAM")
    write_table_header(["Position", "Name", "Background", "LinkedIn", "Key Experience"])
    for i, person in enumerate(report.get("leadership_team", [])):
        write_data_row([
            person.get("position"),
            person.get("name"),
            person.get("background"),
            person.get("linkedin"),
            person.get("key_experience"),
        ], alt=i % 2 == 1)
    write_kv("Team Size", f"{_safe(report.get('team_size'))} — {_safe(report.get('team_breakdown'))}")
    write_spacer()

    # ── MARKET ANALYSIS ──
    write_section_header("MARKET ANALYSIS")
    write_table_header(["Category", "Current Size", "Projection", "CAGR", "Notes"])
    for i, m in enumerate(report.get("market_analysis", [])):
        write_data_row([
            m.get("category"),
            m.get("current_size"),
            m.get("projection"),
            m.get("cagr"),
            m.get("notes"),
        ], alt=i % 2 == 1)
    write_spacer()

    # ── PRODUCT & POSITIONING ──
    write_section_header("PRODUCT & POSITIONING")
    pp = report.get("product_positioning", {})
    write_kv("Core Value Prop", pp.get("core_value_prop"))
    write_kv("Key Differentiators", pp.get("key_differentiators"))
    write_kv("Target Customers", pp.get("target_customers"))
    write_kv("Pricing Model", pp.get("pricing_model"))
    write_spacer()

    # ── TRACTION METRICS ──
    write_section_header("TRACTION METRICS")
    write_table_header(["Metric", "Current Value", "Significance", "", ""])
    for i, t in enumerate(report.get("traction_metrics", [])):
        write_data_row([
            t.get("metric"),
            t.get("current_value"),
            t.get("significance"),
            None, None,
        ], alt=i % 2 == 1)
    write_spacer()

    # ── COMPETITIVE LANDSCAPE ──
    write_section_header("COMPETITIVE LANDSCAPE")
    write_table_header(["Competitor", "Valuation/Funding", "Focus", "Strengths", "Weaknesses vs Target"])
    for i, c in enumerate(report.get("competitive_landscape", [])):
        write_data_row([
            c.get("competitor"),
            c.get("valuation_funding"),
            c.get("focus"),
            c.get("strengths"),
            c.get("weaknesses_vs_target"),
        ], alt=i % 2 == 1)
    write_spacer()

    # ── FINANCIAL ANALYSIS ──
    write_section_header("FINANCIAL ANALYSIS")
    fa = report.get("financial_analysis", {})
    write_kv("Funding Raised", fa.get("funding_raised"))
    write_kv("Investors", fa.get("investors"))
    write_kv("Revenue Model", fa.get("revenue_model"))
    write_kv("Burn Rate", fa.get("burn_rate"))
    write_kv("Runway", fa.get("runway"))
    write_spacer()

    # ── RISK ASSESSMENT ──
    write_section_header("RISK ASSESSMENT")
    write_table_header(["Risk Category", "Level", "Description", "Mitigation", ""])
    for i, r in enumerate(report.get("risk_assessment", [])):
        write_data_row([
            r.get("risk_category"),
            r.get("level"),
            r.get("description"),
            r.get("mitigation"),
            None,
        ], alt=i % 2 == 1)
    write_spacer()

    # ── INVESTMENT RECOMMENDATION ──
    write_section_header("INVESTMENT RECOMMENDATION")
    ir = report.get("investment_recommendation", {})
    # Rating gets accent styling
    c1 = ws.cell(row=row, column=1, value="Overall Rating")
    c1.font = label_font
    c1.border = thin_border
    c2 = ws.cell(row=row, column=2, value=_safe(ir.get("overall_rating")))
    c2.font = accent_font
    c2.border = thin_border
    row += 1

    write_kv("Investment Thesis", ir.get("investment_thesis"))
    write_kv("Key Strengths", ir.get("key_strengths"))
    write_kv("Key Concerns", ir.get("key_concerns"))
    write_kv("Valuation Assessment", ir.get("valuation_assessment"))
    write_kv("Recommended Investment", ir.get("recommended_investment"))
    write_kv("Expected Timeline to Exit", ir.get("expected_timeline_to_exit"))
    write_kv("Risk-Adjusted Return", ir.get("risk_adjusted_return"))

    # Print setup
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(path)
    print(f"  📊 XLSX: {path}")
    return path


if __name__ == "__main__":
    import json, sys
    # Test with the reference CSV data structure
    sample = json.load(open(sys.argv[1])) if len(sys.argv) > 1 else {}
    if sample:
        generate_xlsx(sample, "Test_Company")
        generate_csv(sample, "Test_Company")
